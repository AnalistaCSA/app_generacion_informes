import requests
from flask import Flask, request, send_file
from flask_cors import CORS
import zipfile
import time
from io import BytesIO
from openpyxl import load_workbook
import urllib3
from openpyxl.drawing.image import Image
import os
import json
import re
import traceback
from tempfile import NamedTemporaryFile
from flask import jsonify
import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

API_URL = "https://five.epicollect.net/api/export/entries/ups-colsof?form_ref=055828c4bea94d6fa8a93f8b98adb69f_6a563f807e24e"

with open("backend/data/tecnicos.json", "r", encoding="utf-8") as f:
    tecnicos = json.load(f)

headers = {
    "User-Agent": "Mozilla/5.0"
}

cache_datos = []
ultima_actualizacion = 0

def obtener_datos():

    """
    Consulta la API de EpiCollect y obtiene todos los registros
    disponibles del formulario de instalación de UPS.

    La función implementa paginación automática y cache local
    para reducir el número de consultas realizadas a la API.

    Returns:
        list: Lista de registros obtenidos desde EpiCollect.
    """

    global cache_datos
    global ultima_actualizacion

    # cache 10 minutos
    if time.time() - ultima_actualizacion < 600 and cache_datos:
        print("Usando cache", flush=True)
        return cache_datos

    try:

        todos = []

        # =========================
        # PRIMERA PAGINA
        # =========================

        response = requests.get(
            f"{API_URL}&page=1",
            headers=headers,
            verify=False,
            timeout=30
        )

        print("STATUS PAGINA 1:", response.status_code, flush=True)

        if response.status_code != 200:

            print("ERROR PAGINA 1", flush=True)

            if cache_datos:
                return cache_datos

            return []

        data = response.json()

        total_paginas = data.get("meta", {}).get("last_page", 1)

        print(f"TOTAL PAGINAS: {total_paginas}", flush=True)

        entries = data.get("data", {}).get("entries", [])

        todos.extend(entries)

        # =========================
        # DEMAS PAGINAS
        # =========================

        for page in range(2, total_paginas + 1):

            pagina_cargada = False
            intentos = 0

            while not pagina_cargada and intentos < 10:

                try:

                    print(f"Consultando página {page}", flush=True)

                    response = requests.get(
                        f"{API_URL}&page={page}",
                        headers=headers,
                        verify=False,
                        timeout=30
                    )

                    print(f"STATUS PAGINA {page}: {response.status_code}", flush=True)

                    # RATE LIMIT
                    if response.status_code == 429:

                        intentos += 1

                        print(
                            f"RATE LIMIT PAGINA {page} - intento {intentos}",
                            flush=True
                        )

                        time.sleep(20)

                        continue

                    # ERROR GENERAL
                    if response.status_code != 200:

                        print(f"ERROR PAGINA {page}", flush=True)

                        break

                    data = response.json()

                    entries = data.get("data", {}).get("entries", [])

                    print(
                        f"REGISTROS PAGINA {page}: {len(entries)}",
                        flush=True
                    )

                    todos.extend(entries)

                    print(
                        f"TOTAL ACUMULADO: {len(todos)}",
                        flush=True
                    )

                    pagina_cargada = True

                    time.sleep(5)

                except Exception as e:

                    intentos += 1

                    print(
                        f"ERROR PAGINA {page} - intento {intentos}: {e}",
                        flush=True
                    )

                    time.sleep(10)

        cache_datos = todos
        ultima_actualizacion = time.time()

        print(f"TOTAL REGISTROS FINAL: {len(todos)}", flush=True)

        return todos

    except Exception as e:

        print("ERROR GENERAL:", str(e), flush=True)

        if cache_datos:
            return cache_datos

        return []


def generar_excel(seleccionados=None):

    """
    Genera informes técnicos en formato Excel a partir de los
    registros seleccionados por el usuario.

    También descarga e inserta automáticamente las evidencias
    fotográficas asociadas a cada registro.

    Args:
        seleccionados (list): Lista de identificadores UUID.

    Returns:
        BytesIO: Archivo ZIP con los informes generados.
    """
        
    try:

        print("Consultando API...")

        datos = obtener_datos()
        

        if not datos:
            print("No hay datos, se cancela ejecución")
            return None

        print(f"Registros: {len(datos)}")

        archivos = []

        for item in datos:

            #filtro de seleccion checkbox
            if seleccionados and item.get("ec5_uuid") not in seleccionados:
                continue

            #FUNCION PARA INSERTAR IMAGENES
            def insertar_imagen(ws, item, campo, celda):
                foto1 = item.get(campo)

                if foto1:
                    try:

                        #sacar URL correctamente
                        if isinstance(foto1, list):
                            url = foto1[0].get("file")
                        else:
                            url = foto1

                        #optimizar imagen (más liviana)
                        url = url.replace("entry_original", "entry_small")
                        
                        response = requests.get(foto1, verify=False, timeout=10)

                        if response.status_code == 200:
                            img = Image(BytesIO(response.content))
                            img.width = 358
                            img.height = 340 
                            ws.add_image(img, celda)
                        else:
                            print("No se pudo descargar imagen")
                    except Exception as e:
                        print("Error con imagen:", e)
                else:
                    ws[celda]="N/A"

            print("Procesando: ", item.get("title"))

            wb = load_workbook("formato/formato_informe_ups.xlsx")
            oficinas = load_workbook(r"data/oficinas_ba.xlsx")
            hoja = oficinas.active

            buscar_sban = str(item.get("2_SBAN")).strip()

            n_oficina = ""
            ciudad = ""
            departamento = ""
            direccion = ""

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                if str(fila[0]).strip() == buscar_sban:
                    n_oficina = fila[1]
                    ciudad = fila[2]
                    departamento = fila[3]
                    direccion = fila[4]
                    break
            

            dt_generales = wb["DATOS GENERALES"]
            evi_mantenimiento = wb["EVIDENCIA DEL MANTENIMIENTO"]
            med_entradas_ups = wb["MEDICIONES DE ENTRADA"]
            med_salida_ups = wb["MEDICIONES DE SALIDA"]
            novedades = wb["NOVEDADES"]

            # ***********  Completar datos en HOJA DATOS GENERALES" ***************
            dt_generales["B9"] = n_oficina
            dt_generales["S9"] = item.get("2_SBAN")
            dt_generales["B10"] = item.get("4_FECHA")
            dt_generales["B11"] = ciudad
            dt_generales["B12"] = departamento
            dt_generales["M10"] = item.get("3_NUMERO_DE_RUTINA")
            dt_generales["M11"] = direccion
            dt_generales["M12"] = item.get("5_ENCARGADO")
            dt_generales["S15"] = item.get("8_TIENE_AIRE_ACONDIC")
            dt_generales["U15"] = item.get("9_TEMPERATURA_AMBIEN")

            for tecnico in tecnicos:
                    
                if tecnico.get("id") == item.get("6_CODIGO_DEL_TECNICO"):
                    dt_generales["C40"] = tecnico.get("nombre")
                    dt_generales["C41"] = tecnico.get("documento")
                    dt_generales["C42"] = tecnico.get("telefono")
                    break

            if item.get("11_EL_EQUIPO_ESTA_EN") == "SI":
                dt_generales["C18"] = "X"
            else:
                dt_generales["F18"] = "X"
            if item.get("12_ACCESIBILIDAD_A_E") == "SI":
                dt_generales["C19"] = "X"
            else:
                dt_generales["F19"] = "X"
            if item.get("13_ESTADO_DE_CARCASA") == "SI":
                dt_generales["C20"] = "X"
            else:
                dt_generales["F20"] = "X"
            if item.get("14_ALARMAS_ACTIVAS") == "SI":
                dt_generales["C21"] = "X"
                dt_generales["C22"] = item.get("15_INDIQUE_CODIGO_AL")
            else:
                dt_generales["F21"] = "X"
            if item.get("16_CONEXIONES_DE_ENT") == "SI":
                dt_generales["M18"] = "X"
            else:
                dt_generales["O18"] = "X"
            if item.get("17_CONEXIONES_DE_SAL") == "SI":
                dt_generales["M19"] = "X"
            else:
                dt_generales["O19"] = "X"
            if item.get("18_EQUIPO_EN_GESTION") == "SI":
                dt_generales["M20"] = "X"
            else:
                dt_generales["O20"] = "X"
            if item.get("19_CONDUCTORES_EN_BU") == "SI":
                dt_generales["M21"] = "X"
            else:
                dt_generales["O21"] = "X"
            if item.get("20_CONEXION_DE_BATER") == "SI":
                dt_generales["S18"] = "X"
            else:
                dt_generales["U18"] = "X"
            if item.get("21_CONEXION_A_TIERRA") == "SI":
                dt_generales["S19"] = "X"
            else:
                dt_generales["U19"] = "X"
            if item.get("22_TIENE_PATCH_CORD_") == "SI":
                dt_generales["S20"] = "X"
            else:
                dt_generales["U20"] = "X"
            dt_generales["S21"] = item.get("23_NUMERO_DE_BANCO_D")

            dt_generales["C24"] = item.get("25_TIENE_TARJETA_SNM")
            dt_generales["H24"] = item.get("26_NUMERO_DE_SERIE_D")
            dt_generales["N24"] = item.get("27_NUMERO_DE_MAC")
            dt_generales["S24"] = item.get("28_CONFIGURACION_IP_")
            dt_generales["J15"] = item.get("30_CAPACIDAD_DE_UPS")
            dt_generales["B15"] = item.get("31_MARCA_UPS")
            dt_generales["F15"] = item.get("32_MODELO_UPS")
            dt_generales["M15"] = item.get("33_NUMERO_DE_FASES")
            dt_generales["P15"] = item.get("34_NUMERO_DE_SERIE_U")

            if item.get("35_TIPO_DE_UPS") == "MONOFASICA":                
                med_entradas_ups["C2"] = item.get("35_TIPO_DE_UPS")
                
                #VALOR MEDICIONES DE ENTRADA EN HOJA DATOS GENERALES
                dt_generales["D29"] = "N/A"
                dt_generales["F29"] = "N/A"
                dt_generales["H29"] = "N/A" 
                dt_generales["D30"] = item.get("38_DIGITE_VALOR_TENS")
                dt_generales["F30"] = "N/A"
                dt_generales["H30"] = "N/A"
                dt_generales["C31"] = item.get("40_DIGITE_VALOR_MEDI")
                dt_generales["L29"] = item.get("42_DIGITE_EL_VALOR_M")
                dt_generales["N29"] = "N/A"
                dt_generales["Q29"] = "N/A"
                dt_generales["S29"] = item.get("44_DIGITAR_EL_VALOR_")
                dt_generales["U29"] = item.get("46_DIGITAR_VALOR_ENT")

                #FOTOS MEDICONES DE ENTRADA EN HOJA MEDICIONES DE ENTRADA
                insertar_imagen(med_entradas_ups, item, "37_FOTO_TENSION_ENTR", "A4")
                med_entradas_ups["C4"] = "N/A"
                med_entradas_ups["E4"] = "N/A"
                med_entradas_ups["A22"] = "N/A"
                med_entradas_ups["C22"] = "N/A"
                med_entradas_ups["E22"] = "N/A"
                insertar_imagen(med_entradas_ups, item, "39_FOTO_TENSION_ENTR", "A40")
                insertar_imagen(med_entradas_ups, item, "41_FOTO_CORRIENTE_EN", "C40")
                med_entradas_ups["E40"] = "N/A"
                med_entradas_ups["A58"] = "N/A"
                insertar_imagen(med_entradas_ups, item, "43_FOTO_CORRIENTE_EN", "C58")
                insertar_imagen(med_entradas_ups, item, "45_FOTO_CORREINTE_EN", "E58")
                med_entradas_ups["A86"] = item.get("47_OBERVACIONES_ENTR")

                #VALOR MEDICIONES DE SALIDA EN HOJA DATOS GENERALES
                dt_generales["D34"] = item.get("49_DIGITE_VALOR_MEDI")
                dt_generales["D33"] = "N/A"
                dt_generales["F33"] = "N/A"
                dt_generales["H33"] = "N/A"
                dt_generales["F34"] = "N/A"
                dt_generales["H34"] = "N/A"
                dt_generales["M33"] = item.get("51_DIGITE_VALOR_MEDI")
                dt_generales["L30"] = item.get("53_DIGITE_EL_VALOR_M")
                dt_generales["N30"] = "N/A"
                dt_generales["Q30"] = "N/A"
                dt_generales["S30"] = item.get("55_DIGITAR_EL_VALOR_")
                dt_generales["U30"] = item.get("57_DIGITAR_VALOR_SAL")

                #FOTOS MEDICONES DE SALIDA EN HOJA MEDICIONES DE SALIDA
                insertar_imagen(med_salida_ups, item, "48_FOTO_TENSION_SALI", "A4")
                med_salida_ups["C4"] = "N/A"
                med_salida_ups["E4"] = "N/A"
                med_salida_ups["A22"] = "N/A"
                med_salida_ups["C22"] = "N/A"
                med_salida_ups["E22"] = "N/A"
                insertar_imagen(med_salida_ups, item, "50_FOTO_TENSION_SALI", "A40")
                insertar_imagen(med_salida_ups, item, "52_FOTO_CORRIENTE_SA", "C40")
                med_salida_ups["E40"] = "N/A"
                med_salida_ups["A58"] = "N/A"
                insertar_imagen(med_salida_ups, item, "54_FOTO_CORRIENTE_SA", "C58")
                insertar_imagen(med_salida_ups, item, "56_FOTO_CORRIENTE_SA", "E58")
                med_salida_ups["A109"] = item.get("58_OBERVACIONES_SALI")

            elif item.get("35_TIPO_DE_UPS") == "BIFASICA":
                med_entradas_ups["C2"] = item.get("35_TIPO_DE_UPS")
                #VALOR MEDICIONES DE ENTRADA EN HOJA DATOS GENERALES
                dt_generales["D29"] = item.get("62_DIGITE_MEDIDA_TEN")
                dt_generales["F29"] = "N/A"
                dt_generales["H29"] = "N/A" 
                dt_generales["L29"] = item.get("64_DIGITE_VALOR_CORR")
                dt_generales["N29"] = item.get("66_DIGITE_VALOR_CORR")
                dt_generales["Q29"] = "N/A"
                dt_generales["D30"] = "N/A"
                dt_generales["F30"] = "N/A"
                dt_generales["H30"] = "N/A"
                dt_generales["C31"] = "N/A"
                dt_generales["S29"] = "N/A"
                dt_generales["U29"] = item.get("68_DIGITAR_VALOR_ENT")

                #FOTOS MEDICONES DE ENTRADA EN HOJA MEDICIONES DE ENTRADA
                med_entradas_ups["A4"] = "N/A"
                med_entradas_ups["C4"] = "N/A"
                med_entradas_ups["E4"] = "N/A"
                insertar_imagen(med_entradas_ups, item, "61_FOTO_TENSION_ENTR", "A22")
                med_entradas_ups["C22"] = "N/A"
                med_entradas_ups["E22"] = "N/A"
                med_entradas_ups["A40"] = "N/A"                
                insertar_imagen(med_entradas_ups, item, "63_FOTO_CORRIENTE_EN", "C40")
                insertar_imagen(med_entradas_ups, item, "65_FOTO_CORRIENTE_EN", "E40")
                med_entradas_ups["A58"] = "N/A"
                med_entradas_ups["C58"] = "N/A"
                insertar_imagen(med_entradas_ups, item, "67_FOTO_CORREINTE_EN", "E58")
                med_entradas_ups["A86"] = item.get("69_OBERVACIONES_ENTR")

                #VALOR MEDICIONES DE SALIDA EN HOJA DATOS GENERALES
                dt_generales["D34"] = item.get("71_DIGITAR_VALOR_MED")
                dt_generales["F34"] = item.get("73_DIGITAR_VALOR_MED")
                dt_generales["H34"] = "N/A"
                dt_generales["D33"] = item.get("75_DIGITE_MEDIDA_TEN")
                dt_generales["F33"] = "N/A"
                dt_generales["H33"] = "N/A"
                dt_generales["M33"] = item.get("77_DIGITE_VALOR_MEDI")
                dt_generales["L30"] = item.get("79_DIGITE_VALOR_CORR")
                dt_generales["N30"] = item.get("81_DIGITE_VALOR_CORR")
                dt_generales["Q30"] = "N/A"
                dt_generales["S30"] = item.get("83_DIGITAR_EL_VALOR_")
                dt_generales["U30"] = item.get("85_DIGITAR_VALOR_SAL")

                #FOTOS MEDICONES DE SALIDA EN HOJA MEDICIONES DE SALIDA
                insertar_imagen(med_salida_ups, item, "70_FOTO_TENSION_DE_S", "A4")
                insertar_imagen(med_salida_ups, item, "72_FOTO_TENSION_DE_S", "C4")
                med_salida_ups["E4"] = "N/A"
                insertar_imagen(med_salida_ups, item, "74_FOTO_TENSION_SALI", "A22")
                med_salida_ups["C22"] = "N/A"
                med_salida_ups["E22"] = "N/A"
                insertar_imagen(med_salida_ups, item, "76_FOTO_TENSION_SALI", "A40")
                insertar_imagen(med_salida_ups, item, "78_FOTO_CORRIENTE_SA", "C40")
                insertar_imagen(med_salida_ups, item, "80_FOTO_CORRIENTE_SA", "E40")
                med_salida_ups["A58"] = "N/A"
                insertar_imagen(med_salida_ups, item, "82_FOTO_CORRIENTE_SA", "C58")
                insertar_imagen(med_salida_ups, item, "84_FOTO_CORREINTE_SA", "E58")
                med_salida_ups["A109"] = item.get("86_OBERVACIONES_SALI")

            else: #TRIFASICA
                med_entradas_ups["C2"] = item.get("35_TIPO_DE_UPS")

                #VALOR MEDICIONES DE ENTRADA EN HOJA DATOS GENERALES
                dt_generales["D30"] = item.get("90_DIGITAR_VALOR_MED")
                dt_generales["F30"] = item.get("92_DIGITAR_VALOR_MED")
                dt_generales["H30"] = item.get("94_DIGITAR_VALOR_MED")
                dt_generales["D29"] = item.get("96_DIGITE_MEDIDA_TEN")
                dt_generales["F29"] = item.get("98_DIGITE_MEDIDA_TEN")
                dt_generales["H29"] = item.get("100_DIGITE_MEDIDA_TE")
                dt_generales["L29"] = item.get("104_DIGITE_VALOR_COR")
                dt_generales["N29"] = item.get("106_DIGITE_VALOR_COR")
                dt_generales["Q29"] = item.get("108_DIGITE_VALOR_COR")
                dt_generales["C31"] = item.get("102_DIGITE_VALOR_MED")
                dt_generales["S29"] = item.get("110_DIGITAR_EL_VALOR")
                dt_generales["U29"] = item.get("112_DIGITAR_VALOR_EN")

                #FOTOS MEDICONES DE ENTRADA EN HOJA MEDICIONES DE ENTRADA
                insertar_imagen(med_entradas_ups, item, "89_FOTO_TENSION_DE_E", "A4")
                insertar_imagen(med_entradas_ups, item, "91_FOTO_TENSION_DE_E", "C4")
                insertar_imagen(med_entradas_ups, item, "93_FOTO_TENSION_DE_E", "E4")
                insertar_imagen(med_entradas_ups, item, "95_FOTO_TENSION_ENTR", "A22")
                insertar_imagen(med_entradas_ups, item, "97_FOTO_TENSION_ENTR", "C22")
                insertar_imagen(med_entradas_ups, item, "99_FOTO_TENSION_ENTR", "E22")
                insertar_imagen(med_entradas_ups, item, "101_FOTO_TENSION_ENT", "A40")               
                insertar_imagen(med_entradas_ups, item, "103_FOTO_CORRIENTE_E", "C40")
                insertar_imagen(med_entradas_ups, item, "105_FOTO_CORRIENTE_E", "E40")
                insertar_imagen(med_entradas_ups, item, "107_FOTO_CORRIENTE_E", "A58")
                insertar_imagen(med_entradas_ups, item, "109_FOTO_CORRIENTE_E", "C58")
                insertar_imagen(med_entradas_ups, item, "111_FOTO_CORREINTE_E", "E58")
                med_entradas_ups["A86"] = item.get("113_OBERVACIONES_ENT")

                #VALOR MEDICIONES DE SALIDA EN HOJA DATOS GENERALES
                dt_generales["D34"] = item.get("115_DIGITAR_VALOR_ME")
                dt_generales["F34"] = item.get("117_DIGITAR_VALOR_ME")
                dt_generales["H34"] = item.get("119_DIGITAR_VALOR_ME")
                dt_generales["D33"] = item.get("121_DIGITE_MEDIDA_TE")
                dt_generales["F33"] = item.get("123_DIGITE_MEDIDA_TE")
                dt_generales["H33"] = item.get("125_DIGITE_MEDIDA_TE")
                dt_generales["M33"] = item.get("127_DIGITE_VALOR_MED")
                dt_generales["L30"] = item.get("129_DIGITE_VALOR_COR")
                dt_generales["N30"] = item.get("131_DIGITE_VALOR_COR")
                dt_generales["Q30"] = item.get("133_DIGITE_VALOR_COR")
                dt_generales["S30"] = item.get("135_DIGITAR_EL_VALOR")
                dt_generales["U30"] = item.get("137_DIGITAR_VALOR_SA")

                #FOTOS MEDICONES DE SALIDA EN HOJA MEDICIONES DE SALIDA
                insertar_imagen(med_salida_ups, item, "114_FOTO_TENSION_DE_", "A4")
                insertar_imagen(med_salida_ups, item, "116_FOTO_TENSION_DE_", "C4")
                insertar_imagen(med_salida_ups, item, "118_FOTO_TENSION_DE_", "E4")
                insertar_imagen(med_salida_ups, item, "120_FOTO_TENSION_SAL", "A22")
                insertar_imagen(med_salida_ups, item, "122_FOTO_TENSION_SAL", "C22")
                insertar_imagen(med_salida_ups, item, "124_FOTO_TENSION_SAL", "E22")
                insertar_imagen(med_salida_ups, item, "126_FOTO_TENSION_SAL", "A40")
                insertar_imagen(med_salida_ups, item, "128_FOTO_CORRIENTE_S", "C40")
                insertar_imagen(med_salida_ups, item, "130_FOTO_CORRIENTE_S", "E40")
                insertar_imagen(med_salida_ups, item, "132_FOTO_CORRIENTE_S", "A58")
                insertar_imagen(med_salida_ups, item, "134_FOTO_CORRIENTE_S", "C58")
                insertar_imagen(med_salida_ups, item, "136_FOTO_CORREINTE_S", "E58")
                med_salida_ups["A109"] = item.get("138_OBERVACIONES_SAL")

            if item.get("139_LAS_LECTURAS_COR") == "SI":
                dt_generales["P34"] = "X"
            else:
                dt_generales["U34"] = "X"

            #PRUEBAS Y SOPORTE
            dt_generales["E36"] = item.get("141_TENSION_INICIAL_")
            dt_generales["M36"] = item.get("143_TENSION_EN_SOPOR")
            dt_generales["S36"] = item.get("145_TENSIN_CARGA_VDC")
            med_salida_ups["F100"] = item.get("147_REFERENCIA_DE_BA")
            med_salida_ups["F102"] = item.get("148_CANTIDAD_DE_BATE")
            med_salida_ups["F104"] = item.get("149_HORA_INICIO_PRUE")
            med_salida_ups["F105"] = item.get("150_HORA_FINALIZACIO")
            dt_generales["C37"] = item.get("152_CONCLUSIONES_DE_")

            #FOTOS PRUEBAS Y SOPORTE
            insertar_imagen(med_salida_ups, item, "142_TENSION_DE_BATER","A76")
            insertar_imagen(med_salida_ups, item, "144_TENSION_BATERIAS","C76")
            insertar_imagen(med_salida_ups, item, "146_TENSIN_CARGA_VDC","E76")

            #REGISTRO Y EVIDENCIA FOTOGRAFICA
            insertar_imagen(evi_mantenimiento, item, "154_Panoramica_Ubica","A4")
            insertar_imagen(evi_mantenimiento, item, "155_Vista_Frontal_de","B4")
            insertar_imagen(evi_mantenimiento, item, "156_Placa_con_Serial","C4")
            insertar_imagen(evi_mantenimiento, item, "157_Vista_Interna_Ge","D4")

            insertar_imagen(evi_mantenimiento, item, "158_Panoramica_Ubica","A23")
            insertar_imagen(evi_mantenimiento, item, "159_Vista_Frontal_de","B23")
            insertar_imagen(evi_mantenimiento, item, "160_Placa_con_Activo","C23")
            insertar_imagen(evi_mantenimiento, item, "161_Vista_Interna_Ge","D23")

            insertar_imagen(evi_mantenimiento, item, "162_DPS_de_Salida_UP","A41")
            insertar_imagen(evi_mantenimiento, item, "163_Reemplazo_de_Ven","B41")
            insertar_imagen(evi_mantenimiento, item, "164_Display_de_la_UP","C41")
            insertar_imagen(evi_mantenimiento, item, "165_Display_de_la_UP","D41")

            insertar_imagen(evi_mantenimiento, item, "166_Placa_con_Serial","A59")
            insertar_imagen(evi_mantenimiento, item, "167_Actualizacion_Ve","B59")
            insertar_imagen(evi_mantenimiento, item, "168_Evidencia_de_Ges","C59")
            insertar_imagen(evi_mantenimiento, item, "169_Evidencia_de_Ges","D59")

            insertar_imagen(evi_mantenimiento, item, "170_Panoramica_Stick","A77")
            insertar_imagen(evi_mantenimiento, item, "171_Placa_de_Manteni","B77")
            insertar_imagen(evi_mantenimiento, item, "172_Marquilla_de_Ide","C77")

            evi_mantenimiento["A94"] = item.get("173_OBSERVACIONES")

            #novedades
            insertar_imagen(novedades, item, "175_NOVEDAD_1","A4")
            insertar_imagen(novedades, item, "176_NOVEDAD_2","C4")
            insertar_imagen(novedades, item, "177_NOVEDAD_3","E4")

            if item.get("178_NOVEDADES_AMBIEN") == "OTRA":
                novedades["D21"] = item.get("179_OTRAS_NOVEDADES_")
            else:
                novedades["D21"] = item.get("178_NOVEDADES_AMBIEN")

            if item.get("180_NOVEDADES_DE_UPS") == "OTRA":
                novedades["D22"] = item.get("181_OTRAS_NOVEDADES_")
            else:
                novedades["D22"] = item.get("180_NOVEDADES_DE_UPS")

            #Generacion de archivos
            sban = item.get("2_SBAN", "sin_sban")
            equipo = item.get("34_CAPACIDAD_DE_UPS", "sin_equipo")
            sn = item.get("38_NUMERO_DE_SERIE_U", "sin_sn")

            nombre_archivo = f"informe_mantenimiento_UPS_SBAN_{sban}_{equipo}_KVA_SN_{sn}.xlsx"

            # limpiar TODO lo problemático
            nombre_archivo = re.sub(r'[^\w\-.]', '_', nombre_archivo)

            # asegurar extensión correcta
            if not nombre_archivo.lower().endswith(".xlsx"):
                nombre_archivo += ".xlsx"

            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                wb.save(tmp.name)

                with open(tmp.name, "rb") as f:
                    contenido = f.read()

            archivos.append((nombre_archivo, contenido))

            print(f"Generado: {nombre_archivo}")

        # varios archivos
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for nombre, archivo in archivos:
                
                if isinstance(archivo, BytesIO):
                    archivo.seek(0)
                    zipf.writestr(nombre, archivo.read())
                else:
                    zipf.writestr(nombre, archivo)

        zip_buffer.seek(0)

        return zip_buffer
    
    except Exception as e:

        print("=========== ERROR EN GENERAR_EXCEL ===========")
        print(traceback.format_exc())

        raise e

def obtener_datos_dashboard():

    datos = obtener_datos()

    datos_livianos = []

    oficinas = load_workbook(r"data/oficinas_ba.xlsx")
    hoja = oficinas.active

    buscar_sban = str(item.get("2_SBAN")).strip()

    n_oficina = ""
    ciudad = ""
    departamento = ""
    direccion = ""

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if str(fila[0]).strip() == buscar_sban:
            n_oficina = fila[1]
            ciudad = fila[2]
            departamento = fila[3]
            direccion = fila[4]
            break

    for item in datos:

        datos_livianos.append({
            "ec5_uuid": item.get("ec5_uuid"),
            "title": item.get("title"),
            "created_at": item.get("created_at"),
            "6_CIUDAD": ciudad,
            "7_DEPARTAMENTO": departamento,
            "34_CAPACIDAD_DE_UPS": item.get("30_CAPACIDAD_DE_UPS"),
            "38_NUMERO_DE_SERIE_U": item.get("34_NUMERO_DE_SERIE_U"),
            "10_CODIGO_DEL_TECNIC": item.get("6_CODIGO_DEL_TECNICO"),
            "3_NOMBRE_OFICINA": n_oficina,
            "8_DIRECCION": direccion,
            "2_SBAN": item.get("2_SBAN")
        })

    return datos_livianos

# ============================
# RUTAS DEL MODULO COLSOF
# ============================

def registrar_rutas(app):

    @app.route("/colsof/datos", methods=["GET"])
    def datos_colsof():

        return jsonify(obtener_datos_dashboard())


    @app.route("/colsof/generar", methods=["POST"])
    def generar_colsof():

        try:

            data = request.json
            seleccionados = data.get("ids", [])

            archivo = generar_excel(seleccionados)

            if not archivo:
                return {"error": "No se generaron archivos"}, 500

            if isinstance(archivo, tuple):

                nombre, buffer = archivo

                return send_file(
                    BytesIO(buffer),
                    as_attachment=True,
                    download_name=nombre,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            return send_file(
                archivo,
                as_attachment=True,
                download_name="informes.zip",
                mimetype="application/zip"
            )

        except Exception as e:

            print(traceback.format_exc())

            return {"error": str(e)}, 500