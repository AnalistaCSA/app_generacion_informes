import requests
from flask import Flask, request, send_file
from flask_cors import CORS
import zipfile
import time
from io import BytesIO
from openpyxl import load_workbook
import urllib3
from openpyxl.drawing.image import Image
import os
import json
import re
import traceback
from tempfile import NamedTemporaryFile
from flask import jsonify
import requests

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

API_URL = "https://five.epicollect.net/api/export/entries/csa-ups-instalacion?form_ref=fff4776480684a35b8765ec74e7c14f8_69c54ba08a99d"

with open("backend/data/tecnicos.json", "r", encoding="utf-8") as f:
    tecnicos = json.load(f)

headers = {
    "User-Agent": "Mozilla/5.0"
}

cache_datos = []
ultima_actualizacion = 0

def obtener_datos():

    """
    Consulta la API de EpiCollect y obtiene todos los registros
    disponibles del formulario de instalación de UPS.

    La función implementa paginación automática y cache local
    para reducir el número de consultas realizadas a la API.

    Returns:
        list: Lista de registros obtenidos desde EpiCollect.
    """

    global cache_datos
    global ultima_actualizacion

    # cache 10 minutos
    if time.time() - ultima_actualizacion < 600 and cache_datos:
        print("Usando cache", flush=True)
        return cache_datos

    try:

        todos = []

        # =========================
        # PRIMERA PAGINA
        # =========================

        response = requests.get(
            f"{API_URL}&page=1",
            headers=headers,
            verify=False,
            timeout=30
        )

        print("STATUS PAGINA 1:", response.status_code, flush=True)

        if response.status_code != 200:

            print("ERROR PAGINA 1", flush=True)

            if cache_datos:
                return cache_datos

            return []

        data = response.json()

        total_paginas = data.get("meta", {}).get("last_page", 1)

        print(f"TOTAL PAGINAS: {total_paginas}", flush=True)

        entries = data.get("data", {}).get("entries", [])

        todos.extend(entries)

        # =========================
        # DEMAS PAGINAS
        # =========================

        for page in range(2, total_paginas + 1):

            pagina_cargada = False
            intentos = 0

            while not pagina_cargada and intentos < 10:

                try:

                    print(f"Consultando página {page}", flush=True)

                    response = requests.get(
                        f"{API_URL}&page={page}",
                        headers=headers,
                        verify=False,
                        timeout=30
                    )

                    print(f"STATUS PAGINA {page}: {response.status_code}", flush=True)

                    # RATE LIMIT
                    if response.status_code == 429:

                        intentos += 1

                        print(
                            f"RATE LIMIT PAGINA {page} - intento {intentos}",
                            flush=True
                        )

                        time.sleep(20)

                        continue

                    # ERROR GENERAL
                    if response.status_code != 200:

                        print(f"ERROR PAGINA {page}", flush=True)

                        break

                    data = response.json()

                    entries = data.get("data", {}).get("entries", [])

                    print(
                        f"REGISTROS PAGINA {page}: {len(entries)}",
                        flush=True
                    )

                    todos.extend(entries)

                    print(
                        f"TOTAL ACUMULADO: {len(todos)}",
                        flush=True
                    )

                    pagina_cargada = True

                    time.sleep(5)

                except Exception as e:

                    intentos += 1

                    print(
                        f"ERROR PAGINA {page} - intento {intentos}: {e}",
                        flush=True
                    )

                    time.sleep(10)

        cache_datos = todos
        ultima_actualizacion = time.time()

        print(f"TOTAL REGISTROS FINAL: {len(todos)}", flush=True)

        return todos

    except Exception as e:

        print("ERROR GENERAL:", str(e), flush=True)

        if cache_datos:
            return cache_datos

        return []


def generar_excel(seleccionados=None):

    """
    Genera informes técnicos en formato Excel a partir de los
    registros seleccionados por el usuario.

    También descarga e inserta automáticamente las evidencias
    fotográficas asociadas a cada registro.

    Args:
        seleccionados (list): Lista de identificadores UUID.

    Returns:
        BytesIO: Archivo ZIP con los informes generados.
    """
        
    try:

        print("Consultando API...")

        datos = obtener_datos()
        

        if not datos:
            print("No hay datos, se cancela ejecución")
            return None

        print(f"Registros: {len(datos)}")

        archivos = []

        for item in datos:

            #filtro de seleccion checkbox
            if seleccionados and item.get("ec5_uuid") not in seleccionados:
                continue

            #FUNCION PARA INSERTAR IMAGENES
            def insertar_imagen(ws, item, campo, celda):
                foto1 = item.get(campo)

                if foto1:
                    try:

                        #sacar URL correctamente
                        if isinstance(foto1, list):
                            url = foto1[0].get("file")
                        else:
                            url = foto1

                        #optimizar imagen (más liviana)
                        url = url.replace("entry_original", "entry_small")
                        
                        response = requests.get(foto1, verify=False, timeout=10)

                        if response.status_code == 200:
                            img = Image(BytesIO(response.content))
                            img.width = 358
                            img.height = 340 
                            ws.add_image(img, celda)
                        else:
                            print("No se pudo descargar imagen")
                    except Exception as e:
                        print("Error con imagen:", e)
                else:
                    ws[celda]="N/A"

            print("Procesando: ", item.get("title"))

            wb = load_workbook("formato/formato_informe_instalacion_ups.xlsx")

            dt_generales = wb["DATOS GENERALES"]
            evi_instalacion = wb["EVIDENCIA DE LA INSTALACIÓN"]
            med_entradas_ups = wb["MEDICIONES ENTRADA DE UPS"]
            med_salida_ups = wb["MEDICIONES SALIDA DE UPS"]
            display_ups = wb["DISPLAY DE LA UPS"]
            baterias = wb["BATERIAS"]
            novedades = wb["NOVEDADES"]

            # ***********  escribir datos HOJA DATOS GENERALES" ***************
            dt_generales["B9"] = item.get("2_ID_SEDE")
            dt_generales["M9"] = f'{item.get("67_MARCA_UPS")} {item.get("68_MODELO_UPS")}'
            dt_generales["B10"] = item.get("created_at", "").split("T")[0]
            dt_generales["M10"] = item.get("6_DIRECCION")
            dt_generales["B11"] = item.get("4_REGIONAL")
            dt_generales["M11"] = item.get("7_CIUDAD")
            dt_generales["B12"] = item.get("5_NOMBRE_SEDE")
            dt_generales["M12"] = item.get("8_DEPARTAMENTO")
            dt_generales["B13"] = item.get("9_COORDINADOR__ENCAR")
            dt_generales["M13"] = item.get("10_TELEFONO_DE_ENCAR")

            for tecnico in tecnicos:
        
                if tecnico.get("id") == item.get("11_CODIGO_TECNICO"):
                    dt_generales["C66"] = tecnico.get("nombre")
                    dt_generales["C67"] = tecnico.get("documento")
                    dt_generales["C68"] = tecnico.get("telefono")
                    break

            #DATOS GENERALES - DATOS DE LA UPS INSTALADA
            dt_generales["B16"] = item.get("67_MARCA_UPS")
            dt_generales["F16"] = item.get("68_MODELO_UPS")
            dt_generales["J16"] = item.get("66_CAPACIDAD_UPS_KVA")
            dt_generales["M16"] = item.get("33_NUMERO_FASES_ENTR")
            dt_generales["P16"] = item.get("69_NUMERO_DE_SERIE_D")
            dt_generales["S16"] = item.get("13_TIENE_AIRE_ACONDI")
            dt_generales["U16"] = item.get("14_TEMPERATURA_AMBIE")
            dt_generales["B17"] = item.get("41_TIENE_TARJETA_SNM")
            dt_generales["G17"] = item.get("42_MARCA")
            dt_generales["Q17"] = item.get("43_NUMERO_SERIE_SNMP")
            dt_generales["F18"] = item.get("25_NOMBRE_UBICACIN_E")

            #DATOS GENERALES - ÁREA DE UBICACION Y ESTADO DE CONEXIONES
            if item.get("17_DISTANCIA_AL_LADO")=="SI":
                dt_generales["F21"]="X"
            else: dt_generales["I21"]="X"
            if item.get("18_DISTANCIA_AL_FREN")=="SI":
                dt_generales["F22"]="X"
            else: dt_generales["I22"]="X"
            if item.get("19_DISTANCIA_AL_LADO")=="SI":
                dt_generales["P21"]="X"
            else: dt_generales["T21"]="X"
            if item.get("20_PUERTA_DE_TABLERO")=="SI":
                dt_generales["P22"]="X"
            else: dt_generales["T22"]="X"
            dt_generales["E23"] = item.get("21_OBSERVACIONES_DIS")
            if item.get("24_EQUIPO_EN_CUARTO_")=="SI":
                dt_generales["C26"]="X"
            else: dt_generales["F26"]="X"
            if item.get("16_BUENA_ACCESIBILID")=="SI":
                dt_generales["C27"]="X"
            else: dt_generales["F27"]="X"
            if item.get("23_BUEN_ESTADO_DE_CA")=="SI":
                dt_generales["C28"]="X"
            else: dt_generales["F28"]="X"
            if item.get("26_ALARMAS_ACTIVAS")=="SI":
                dt_generales["C29"]="X"
                dt_generales["C30"]=item.get("27_INDIQUE_CODIGO_DE")
            else: dt_generales["F29"]="X"
            if item.get("29_CONEXION_ENTRADAS")=="SI":
                dt_generales["M26"]="X"
            else: dt_generales["O26"]="X"
            if item.get("30_CONEXION_SALIDAS_")=="SI":
                dt_generales["M27"]="X"
            else: dt_generales["O27"]="X"
            if item.get("31_CONEXION_DE_BATER")=="SI":
                dt_generales["S26"]="X"
            else: dt_generales["U26"]="X"
            if item.get("32_CONEXION_A_TIERRA")=="SI":
                dt_generales["S27"]="X"
            else: dt_generales["U27"]="X"
            if item.get("44_EQUIPO_EN_GESTION")=="SI":
                dt_generales["S28"]="X"
            else: dt_generales["U28"]="X"
            if item.get("45_PATCH_CORD_CONECT")=="SI":
                dt_generales["M28"]="X"
            else: dt_generales["O28"]="X"
            if item.get("46_CONDUCTORES_EN_BU")=="SI":
                dt_generales["M29"]="X"
            else: dt_generales["O29"]="X"

            #DATOS GENERALES - INSPECCION DE LAS INSTALACIONES
            dt_generales["I32"] = item.get("33_NUMERO_FASES_ENTR")
            dt_generales["D35"] = item.get("34_CANALIZACION_Y_DI")
            dt_generales["J35"] = item.get("35_LONGITUD_CANALIZA")
            dt_generales["R32"] = item.get("36_NUMERO_FASES_SALI")
            dt_generales["O35"] = item.get("37_CANALIZACION_Y_DI")
            dt_generales["T35"] = item.get("38_LONGITUD_CANALIZA")
            dt_generales["D36"] = item.get("39_OBSERVACIONES_INS")

            #DATOS GENERALES - TABLERO ENTRADA Y SALIDA DE UPS
            #Entrda
            dt_generales["D41"] = item.get("51_BREAKER_ENTRADA_U")
            dt_generales["I41"] = item.get("52_TIPO_ENTRADA_BREA")
            dt_generales["K41"] = item.get("53_MARCA_BREAKER_ENT")
            dt_generales["Q41"] = item.get("54_CAPACIDAD_BREAKER")
            dt_generales["V41"] = item.get("55_CAPACIDAD_DE_CORT")
            dt_generales["D42"] = item.get("48_DPS_ENTRADA")
            dt_generales["I42"] = item.get("49_MARCA_DPS")
            dt_generales["Q42"] = item.get("50_REFERENCIA_DPS")
            #Salida
            dt_generales["D44"] = item.get("60_BREAKER_SALIDA_UP")
            dt_generales["I44"] = item.get("61_TIPO_SALIDA_BREAK")
            dt_generales["K44"] = item.get("62_MARCA_BREAKER_SAL")
            dt_generales["Q44"] = item.get("63_CAPACIDAD_BREAKER")
            dt_generales["V44"] = item.get("64_CAPACIDAD_DE_CORT")
            dt_generales["D45"] = item.get("57_DPS_SALIDA")
            dt_generales["I45"] = item.get("58_MARCA_DPS")
            dt_generales["Q45"] = item.get("59_REFERENCIA_DPS")

            #DATOS SEGUN TIPO DE UPS (MONOFASICA, BIFASICA, TRIFASICA)
            # *************** hojas mediciones entradas y salidas ***************
            if item.get("70_TIPO_DE_UPS") == "MONOFASICA":
                #Fotos y datos entrada
                med_entradas_ups["C2"] = item.get("70_TIPO_DE_UPS")
                dt_generales["E33"] = item.get("72_CALIBRE_ENTRADA_F")
                dt_generales["I33"] = item.get("73_CALIBRE_ENTRADA_N")
                dt_generales["K33"] = item.get("74_CALIBRE_ENTRADA_T")
                dt_generales["E34"] = item.get("75_COLOR_MARQUILLADO")
                
                insertar_imagen(med_entradas_ups, item, "78_FOTO_TENSION_ENTR","A4")
                dt_generales["D49"] = item.get("79_DIGITE_VALOR_MEDI")
                insertar_imagen(med_entradas_ups, item, "80_FOTO_TENSION_ENTR", "A40")
                dt_generales["C50"] = item.get("81_DIGITE_VALOR_MEDI")
                insertar_imagen(med_entradas_ups, item, "82_FOTO_CORRIENTE_EN", "C40")
                dt_generales["L48"] = item.get("83_DIGITE_EL_VALOR_M")
                insertar_imagen(med_entradas_ups, item, "84_FOTO_CORRIENTE_EN", "C58")
                dt_generales["S48"] = item.get("85_DIGITAR_EL_VALOR_")
                insertar_imagen(med_entradas_ups, item, "86_FOTO_CORREINTE_EN", "E58")
                dt_generales["U48"] = item.get("87_DIGITAR_VALOR_ENT")
                med_entradas_ups["A86"] = item.get("88_OBERVACIONES_ENTR")

                #Fotos y datos salida
                med_salida_ups["C2"] = item.get("70_TIPO_DE_UPS")
                dt_generales["R33"] = item.get("89_CALIBRE_SALIDA_FA")
                dt_generales["T33"] = item.get("90_CALIBRE_SALIDA_NE")
                dt_generales["V33"] = item.get("91_CALIBRE_SALIDA_TI")
                dt_generales["R34"] = item.get("92_COLOR_MARQUILLADO")

                insertar_imagen(med_salida_ups, item, "95_FOTO_TENSION_SALI", "A4")
                dt_generales["D53"] = item.get("96_DIGITE_VALOR_MEDI")
                insertar_imagen(med_salida_ups, item, "97_FOTO_TENSION_SALI", "A40")
                dt_generales["M52"] = item.get("98_DIGITE_VALOR_MEDI")
                insertar_imagen(med_salida_ups, item, "99_FOTO_CORRIENTE_SA", "C40")
                dt_generales["L49"] = item.get("100_DIGITE_EL_VALOR_")
                insertar_imagen(med_salida_ups, item, "101_FOTO_CORRIENTE_S", "C58")
                dt_generales["S49"] = item.get("102_DIGITAR_EL_VALOR")
                insertar_imagen(med_salida_ups, item, "103_FOTO_CORREINTE_S", "E58")
                dt_generales["U49"] = item.get("104_DIGITAR_VALOR_SA")
                med_salida_ups["A86"] = item.get("105_OBERVACIONES_SAL")
                
                #valores vacios en ups monofasica
                dt_generales["I34"] = "N/A"
                dt_generales["K34"] = "N/A"
                dt_generales["T34"] = "N/A"
                dt_generales["V34"] = "N/A"
                med_entradas_ups["C4"] = "N/A" 
                med_entradas_ups["E4"] = "N/A"
                med_salida_ups["C4"] = "N/A"
                med_salida_ups["E4"] = "N/A"
                med_entradas_ups["A22"] = "N/A"
                med_entradas_ups["C22"] = "N/A"
                med_entradas_ups["E22"] = "N/A"
                med_salida_ups["A22"] = "N/A"
                med_salida_ups["C22"] = "N/A"
                med_salida_ups["E22"] = "N/A"
                med_entradas_ups["E40"] = "N/A"
                med_salida_ups["E40"] = "N/A"
                med_entradas_ups["A58"] = "N/A"
                med_salida_ups["A58"] = "N/A"
                dt_generales["D48"] = "N/A"           
                dt_generales["F48"] = "N/A"
                dt_generales["H48"] = "N/A"
                dt_generales["F49"] = "N/A"
                dt_generales["H49"] = "N/A"
                dt_generales["N48"] = "N/A"
                dt_generales["Q48"] = "N/A"
                dt_generales["D52"] = "N/A"
                dt_generales["F52"] = "N/A"
                dt_generales["H52"] = "N/A"
                dt_generales["F53"] = "N/A"
                dt_generales["H53"] = "N/A"
                dt_generales["N49"] = "N/A"
                dt_generales["Q49"] = "N/A"

            elif item.get("70_TIPO_DE_UPS") == "BIFASICA":
                #Fotos y datos entrada
                med_entradas_ups["C2"] = item.get("70_TIPO_DE_UPS")
                dt_generales["E33"] = item.get("108_CALIBRE_ENTRADA_")
                dt_generales["K33"] = item.get("109_CALIBRE_ENTRADA_")
                dt_generales["E34"] = item.get("110_COLOR_MARQUILLAD")
                dt_generales["I34"] = item.get("111_COLOR_MARQUILLAD")

                insertar_imagen(med_entradas_ups, item, "113_FOTO_TENSION_ENT", "A22")
                dt_generales["D48"] = item.get("114_DIGITE_MEDIDA_TE")
                insertar_imagen(med_entradas_ups, item, "115_FOTO_CORRIENTE_E", "C40")
                dt_generales["L48"] = item.get("116_DIGITE_VALOR_COR")
                insertar_imagen(med_entradas_ups, item, "117_FOTO_CORRIENTE_E", "E40")
                dt_generales["N48"] = item.get("118_DIGITE_VALOR_COR")
                insertar_imagen(med_entradas_ups, item, "119_FOTO_CORREINTE_E", "E58")
                dt_generales["U48"] = item.get("120_DIGITAR_VALOR_EN")
                med_entradas_ups["A86"] = item.get("121_OBERVACIONES_ENT")

                #fotos y datos salida
                med_salida_ups["C2"] = item.get("70_TIPO_DE_UPS")
                dt_generales["R33"] = item.get("122_CALIBRE_SALIDA_F")
                dt_generales["T33"] = item.get("123_CALIBRE_SALIDA_N")
                dt_generales["V33"] = item.get("124_CALIBRE_SALIDA_T")
                dt_generales["R34"] = item.get("125_COLOR_MARQUILLAD")
                dt_generales["T34"] = item.get("126_COLOR_MARQUILLAD")

                insertar_imagen(med_salida_ups, item, "129_FOTO_TENSION_DE_", "A4")
                dt_generales["D53"] = item.get("130_DIGITAR_VALOR_ME")
                insertar_imagen(med_salida_ups, item, "131_FOTO_TENSION_DE_", "C4")
                dt_generales["F53"] = item.get("132_DIGITAR_VALOR_ME")
                insertar_imagen(med_salida_ups, item, "133_FOTO_TENSION_SAL", "A22")
                dt_generales["D52"] = item.get("134_DIGITE_MEDIDA_TE")
                insertar_imagen(med_salida_ups, item, "135_FOTO_TENSION_SAL", "A40")
                dt_generales["M52"] = item.get("136_DIGITE_VALOR_MED")
                insertar_imagen(med_salida_ups, item, "137_FOTO_CORRIENTE_S", "C40")
                dt_generales["L49"] = item.get("138_DIGITE_VALOR_COR")
                insertar_imagen(med_salida_ups, item, "139_FOTO_CORRIENTE_S", "E40")
                dt_generales["N49"] = item.get("140_DIGITE_VALOR_COR")
                insertar_imagen(med_salida_ups, item, "141_FOTO_CORRIENTE_S", "C58")
                dt_generales["S49"] = item.get("142_DIGITAR_EL_VALOR")
                insertar_imagen(med_salida_ups, item, "143_FOTO_CORREINTE_S", "E58")
                dt_generales["U49"] = item.get("144_DIGITAR_VALOR_SA")
                med_salida_ups["A86"] = item.get("145_OBERVACIONES_SAL")

                #valores vacios en ups BIFASICA
                dt_generales["I33"] = "N/A"
                dt_generales["K34"] = "N/A"
                dt_generales["V34"] = "N/A"
                med_entradas_ups["A4"] = "N/A"
                med_entradas_ups["C4"] = "N/A"
                med_entradas_ups["E4"] = "N/A"
                med_entradas_ups["C22"] = "N/A"
                med_entradas_ups["E22"] = "N/A"
                med_entradas_ups["A40"] = "N/A"
                med_entradas_ups["A58"] = "N/A"
                med_entradas_ups["C58"] = "N/A"
                med_salida_ups["E4"] = "N/A"
                dt_generales["H53"] = "N/A"
                dt_generales["F52"] = "N/A"
                dt_generales["H52"] = "N/A"
                med_salida_ups["C22"] = "N/A"
                med_salida_ups["E22"] = "N/A"
                med_salida_ups["A58"] = "N/A"
                dt_generales["Q49"] = "N/A"
                dt_generales["F48"] = "N/A"
                dt_generales["H48"] = "N/A"
                dt_generales["D49"] = "N/A"
                dt_generales["H49"] = "N/A"
                dt_generales["F49"] = "N/A"
                dt_generales["Q48"] = "N/A"
                dt_generales["S48"] = "N/A"
                dt_generales["C50"] = "N/A"

            else:
                #fotos y datos entrada
                med_entradas_ups["C2"] = item.get("70_TIPO_DE_UPS")
                dt_generales["E33"] = item.get("148_CALIBRE_ENTRADA_")
                dt_generales["I33"] = item.get("149_CALIBRE_ENTRADA_")
                dt_generales["K33"] = item.get("150_CALIBRE_ENTRADA_")
                dt_generales["E34"] = item.get("151_COLOR_MARQUILLAD")
                dt_generales["I34"] = item.get("152_COLOR_MARQUILLAD")
                dt_generales["K34"] = item.get("153_COLOR_MARQUILLAD")

                insertar_imagen(med_entradas_ups, item, "156_FOTO_TENSION_DE_", "A4")
                dt_generales["D49"] = item.get("157_DIGITAR_VALOR_ME")
                insertar_imagen(med_entradas_ups, item, "158_FOTO_TENSION_DE_", "C4")
                dt_generales["F49"] = item.get("159_DIGITAR_VALOR_ME")
                insertar_imagen(med_entradas_ups, item, "160_FOTO_TENSION_DE_", "E4")
                dt_generales["H49"] = item.get("161_DIGITAR_VALOR_ME")
                insertar_imagen(med_entradas_ups, item, "162_FOTO_TENSION_ENT", "A22")
                dt_generales["D48"] = item.get("163_DIGITE_MEDIDA_TE")
                insertar_imagen(med_entradas_ups, item, "164_FOTO_TENSION_ENT", "C22")
                dt_generales["F48"] = item.get("165_DIGITE_MEDIDA_TE")
                insertar_imagen(med_entradas_ups, item, "166_FOTO_TENSION_ENT", "E22")
                dt_generales["H48"] = item.get("167_DIGITE_MEDIDA_TE")
                insertar_imagen(med_entradas_ups, item, "168_FOTO_TENSION_ENT", "A40")
                dt_generales["C50"] = item.get("169_DIGITE_VALOR_MED")
                insertar_imagen(med_entradas_ups, item, "170_FOTO_CORRIENTE_E", "C40")
                dt_generales["L48"] = item.get("171_DIGITE_VALOR_COR")
                insertar_imagen(med_entradas_ups, item, "172_FOTO_CORRIENTE_E", "E40")
                dt_generales["N48"] = item.get("173_DIGITE_VALOR_COR")
                insertar_imagen(med_entradas_ups, item, "174_FOTO_CORRIENTE_E", "A58")
                dt_generales["Q48"] = item.get("175_DIGITE_VALOR_COR")
                insertar_imagen(med_entradas_ups, item, "176_FOTO_CORRIENTE_E", "C58")
                dt_generales["S48"] = item.get("177_DIGITAR_EL_VALOR")
                insertar_imagen(med_entradas_ups, item, "178_FOTO_CORREINTE_E", "E58")
                dt_generales["U48"] = item.get("179_DIGITAR_VALOR_EN")
                med_entradas_ups["A86"] = item.get("180_OBERVACIONES_ENT")

                #fotos y datos salida
                med_salida_ups["C2"] = item.get("70_TIPO_DE_UPS")
                dt_generales["R33"] = item.get("181_CALIBRE_SALIDA_F")
                dt_generales["T33"] = item.get("182_CALIBRE_SALIDA_N")
                dt_generales["V33"] = item.get("183_CALIBRE_SALIDA_T")
                dt_generales["R34"] = item.get("184_COLOR_MARQUILLAD")
                dt_generales["T34"] = item.get("185_COLOR_MARQUILLAD")
                dt_generales["V34"] = item.get("186_COLOR_MARQUILLAD")
                insertar_imagen(med_salida_ups, item, "189_FOTO_TENSION_DE_", "A4")
                dt_generales["D53"] = item.get("190_DIGITAR_VALOR_ME")
                insertar_imagen(med_salida_ups, item, "191_FOTO_TENSION_DE_", "C4")
                dt_generales["F53"] = item.get("192_DIGITAR_VALOR_ME")
                insertar_imagen(med_salida_ups, item, "193_FOTO_TENSION_DE_", "E4")
                dt_generales["H53"] = item.get("194_DIGITAR_VALOR_ME")
                insertar_imagen(med_salida_ups, item, "195_FOTO_TENSION_SAL", "A22")
                dt_generales["D52"] = item.get("196_DIGITE_MEDIDA_TE")
                insertar_imagen(med_salida_ups, item, "197_FOTO_TENSION_SAL", "C22")
                dt_generales["F52"] = item.get("198_DIGITE_MEDIDA_TE")
                insertar_imagen(med_salida_ups, item, "199_FOTO_TENSION_SAL", "E22")
                dt_generales["H52"] = item.get("200_DIGITE_MEDIDA_TE")
                insertar_imagen(med_salida_ups, item, "201_FOTO_TENSION_SAL", "A40")
                dt_generales["M52"] = item.get("202_DIGITE_VALOR_MED")
                insertar_imagen(med_salida_ups, item, "203_FOTO_CORRIENTE_S", "C40")
                dt_generales["L49"] = item.get("204_DIGITE_VALOR_COR")
                insertar_imagen(med_salida_ups, item, "205_FOTO_CORRIENTE_S", "E40")
                dt_generales["N49"] = item.get("206_DIGITE_VALOR_COR")
                insertar_imagen(med_salida_ups, item, "207_FOTO_CORRIENTE_S", "A58")
                dt_generales["Q49"] = item.get("208_DIGITE_VALOR_COR")
                insertar_imagen(med_salida_ups, item, "209_FOTO_CORRIENTE_S", "C58")
                dt_generales["S49"] = item.get("210_DIGITAR_EL_VALOR")
                insertar_imagen(med_salida_ups, item, "211_FOTO_CORREINTE_S", "E58")
                dt_generales["U49"] = item.get("212_DIGITAR_VALOR_SA")
                med_salida_ups["A86"] = item.get("213_OBERVACIONES_SAL")
            
            #datos adiconales hoja datos generales
            if item.get("215_LECTURAS_CORRESP") == "SI":
                dt_generales["P53"] = "X"
            else:
                dt_generales["U53"] ="x"

            if item.get("216_QUE_ALIMENTA_LA_") == "PDU":
                dt_generales["B39"] = "X"
            elif item.get("216_QUE_ALIMENTA_LA_") == "TABLERO REGULADO":
                dt_generales["F39"] = "X"
            elif item.get("216_QUE_ALIMENTA_LA_") == "MULTITOMA":
                dt_generales["K39"] = "X"
            elif item.get("216_QUE_ALIMENTA_LA_") == "RACK DIRECTAMENTE":
                dt_generales["O39"] = "X"
            else:
                dt_generales["R39"] = "X"
                dt_generales["U39"] = item.get("217_CUAL_OTRO")

            # *************** Hoja evidencia fografica instalación ***************
            insertar_imagen(evi_instalacion, item, "220_PANORAMICA_UBICA", "A3")
            insertar_imagen(evi_instalacion, item, "221_VISTA_FRONTAL_UP", "B3")
            insertar_imagen(evi_instalacion, item, "222_VISTA_POSTERIOR_", "C3")
            insertar_imagen(evi_instalacion, item, "223_VISTA_PARTE_SUPE", "D3")
            insertar_imagen(evi_instalacion, item, "224_VISTA_LATERAL_IZ", "A21")
            insertar_imagen(evi_instalacion, item, "225_VISTA_LATERAL_DE", "B21")
            insertar_imagen(evi_instalacion, item, "226_PLACA_SERIAL_DE_", "C21")
            insertar_imagen(evi_instalacion, item, "227_PLACA_NUMERO_ENT", "D21")
            insertar_imagen(evi_instalacion, item, "228_FOTO_SERIE_SNMP", "A39")
            insertar_imagen(evi_instalacion, item, "229_FOTO_BREAKER_ENT", "B39")
            insertar_imagen(evi_instalacion, item, "230_FOTO_DPS_ENTRADA", "C39")
            insertar_imagen(evi_instalacion, item, "231_FOTO_BREAKER_SAL", "D39")
            insertar_imagen(evi_instalacion, item, "232_FOTO_DPS_SALIDA", "A57")
            insertar_imagen(evi_instalacion, item, "233_FOTO_SISTEMA_DE_", "B57")
            insertar_imagen(evi_instalacion, item, "234_FOTO_CANALIZACIO", "C57")
            insertar_imagen(evi_instalacion, item, "235_FOTO_CANALIZACIO", "D57")
            insertar_imagen(evi_instalacion, item, "236_FOTO_BARRAJE_EQU", "A75")
            insertar_imagen(evi_instalacion, item, "237_FOTO_DIAGRAMA_UN", "B75")
            insertar_imagen(evi_instalacion, item, "238_FOTO_VISTA_FRONT", "C75")
            insertar_imagen(evi_instalacion, item, "239_FOTO_TABLERO_RED", "D75")
            insertar_imagen(evi_instalacion, item, "240_FOTO_TABLERO_RED", "A93")
            insertar_imagen(evi_instalacion, item, "241_FOTO_AIRE_ACONDI", "B93")
            evi_instalacion["A111"] = item.get("242_OBSERVACIONES_DE")

            # *************** hoja Baterias ***************
            baterias["B3"] = item.get("244_MARCA_BATERIAS")
            baterias["D3"] = item.get("245_REFERENCIA_BATER")
            baterias["F3"] = item.get("246_AMPERAJE_BATERIA")
            baterias["H3"] = item.get("247_CANTIDAD_BATERIA")

            insertar_imagen(baterias, item, "248_FOTO_MARCA_DE_BA", "A5")
            insertar_imagen(baterias, item, "249_FOTO_CARACTERIAS", "C5")
            insertar_imagen(baterias, item, "250_FOTO_FUSIBLES_PR", "E5")
            insertar_imagen(baterias, item, "251_FOTO_TENSION_NOR", "G5")
            insertar_imagen(baterias, item, "253_FOTO_TENSION_SOP", "A23")
            insertar_imagen(baterias, item, "255_FOTO_TENSION_BAT", "C23")
            if item.get("257_OBSERVACION_PRUE") == "OTRA":
                baterias["A43"] = item.get("257_OBSERVACION_PRUE")
            else:
                baterias["A43"] = item.get("257_OBSERVACION_PRUE")

            #datos adiconales baterias en hoja datos generales
            dt_generales["E62"] = item.get("252_DIGITE_TENSION_I")
            dt_generales["M62"] = item.get("254_DIGITE_TENSION_B")
            dt_generales["S62"] = item.get("256_DIGITE_TENSION_B")

            # *************** hoja display ***************
            if item.get("259_EL_DISPLAY_DE_QU") == "TRIFASICA":
                display_ups["C2"] = item.get("259_EL_DISPLAY_DE_QU")
                display_ups["A3"] = "Display Información UPS Acerca de"
                insertar_imagen(display_ups, item, "271_DISPLAY_INFORMAC", "A4")
                display_ups["C3"] = "Display Condición Normal de Operación En Línea"
                insertar_imagen(display_ups, item, "272_DISPLAY_CONDICIO", "C4")
                display_ups["E3"] = "Display Condición Soporte en Modo Baterías"
                insertar_imagen(display_ups, item, "273_DISPLAY_CONDICIO", "E4")
                display_ups["G3"] = "Display Potencia de Salida Pantalla 1/5"
                insertar_imagen(display_ups, item, "274_DISPLAY_POTENCIA", "G4")
                display_ups["A21"] = "Display Datos de Salida Pantalla 2/5"
                insertar_imagen(display_ups, item, "275_DISPLAY_DATOS_DE", "A22")
                display_ups["C21"] = "Display Datos de Batería Pantalla 3/5"
                insertar_imagen(display_ups, item, "276_DISPLAY_DATOS_DE", "C22")
                display_ups["E21"] = "Display Datos de Bypass Pantalla 4/5"
                insertar_imagen(display_ups, item, "277_DISPLAY_DATOS_DE", "E22")
                display_ups["G21"] = "Display Datos de Bypass Pantalla 4/5"
                insertar_imagen(display_ups, item, "278_DISPLAY_DATOS_DE", "G22")
            else:
                display_ups["C2"] = item.get("259_EL_DISPLAY_DE_QU")
                display_ups["A3"] = "Foto Display Operacion En Condicon En Linea"
                insertar_imagen(display_ups, item, "261_FOTO_DISPLAY_OPE", "A4")
                display_ups["C3"] = "Foto Display Condicion Soporte Baterias"
                insertar_imagen(display_ups, item, "262_FOTO_DISPLAY_CON", "C4")
                display_ups["E3"] = "Visualizacion Display Tension Ac De Entrada"
                insertar_imagen(display_ups, item, "263_VISUALIZACION_DI", "E4")
                display_ups["G3"] = "Visualizacion Tension De Baterias En Flotacion"
                insertar_imagen(display_ups, item, "264_VISUALIZACION_TE", "G4")
                display_ups["A21"] = "Visualizacion Tension De Baterias En Flotacion"
                insertar_imagen(display_ups, item, "265_VISUALIZACION_DI", "A22")
                display_ups["C21"] = "Visualizacion Display Frecuencia De Salida"
                insertar_imagen(display_ups, item, "266_VISUALIZACION_DI", "C22")
                display_ups["E21"] = "Visualizacion Display Frecuencia De Salida"
                insertar_imagen(display_ups, item, "267_VISUALIZACION_DI", "E22")
                display_ups["G21"] = "Visualizacion Display Porcentaje De Carga"
                insertar_imagen(display_ups, item, "268_VISUALIZACION_DI", "G22")

            if item.get("268_VISUALIZACION_DI") == "OTRA":
                display_ups["A40"] = item.get("281_OTRA_OBSERVACION")
            else:
                display_ups["A40"] = item.get("280_OBSERVACION_DISP")

            #datos adicionales de display en HOJA DATOS GENERALES
            if item.get("283_FUNCIONAMIENTO_N") == "SI":
                dt_generales["C57"] = "X"
            else:
                dt_generales["F57"] = "X"

            if item.get("284_APAGADO") == "SI":
                dt_generales["C58"] = "X"
            else:
                dt_generales["F58"] = "X"

            if item.get("285_ENCEDIDO") == "SI":
                dt_generales["C59"] = "X"
            else:
                dt_generales["F59"] = "X"
            
            if item.get("286_OBSERVACIONES_DE") == "OTRA":
                dt_generales["I57"] = item.get("287_OTRA_OBSERVACION")
            else:
                dt_generales["I57"] = item.get("286_OBSERVACIONES_DE")

            # *************** hoja novedades ***************
            novedades["C2"] = item.get("70_TIPO_DE_UPS")
            insertar_imagen(novedades, item, "289_FOTO_NOVEDAD_1", "A4")
            insertar_imagen(novedades, item, "290_FOTO_NOVEDAD_2", "C4")
            insertar_imagen(novedades, item, "291_FOTO_NOVEDAD_3", "E4")
            insertar_imagen(novedades, item, "292_FOTO_NOVEDAD_4", "G4")
            insertar_imagen(novedades, item, "293_FOTO_NOVEDAD_5", "A22")

            if item.get("294_NOVEDADES_AMBIEN") == "OTRA":
                novedades["E38"] = item.get("295_OTRA_OBSERVACION")
            else:
                novedades["E38"] = item.get("294_NOVEDADES_AMBIEN")
            
            if item.get("296_NOVEDADES_UPS") == "OTRA":
                novedades["E39"] = item.get("297_OTRA_OBSERVACION")
            else:
                novedades["E39"] = item.get("296_NOVEDADES_UPS")

            #Generacion de archivos
            titulo = item.get("title", "sin_titulo")
            sede = item.get("5_NOMBRE_SEDE", "sin_sede")
            equipo = item.get("66_CAPACIDAD_UPS_KVA")
            sn = item.get("69_NUMERO_DE_SERIE_D")

            nombre_archivo = f"informe_instalacion_UPS_{titulo}_{sede}_{equipo}_KVA_SN_{sn}.xlsx"

            # limpiar TODO lo problemático
            nombre_archivo = re.sub(r'[^\w\-.]', '_', nombre_archivo)

            # asegurar extensión correcta
            if not nombre_archivo.lower().endswith(".xlsx"):
                nombre_archivo += ".xlsx"

            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                wb.save(tmp.name)

                with open(tmp.name, "rb") as f:
                    contenido = f.read()

            archivos.append((nombre_archivo, contenido))

            print(f"Generado: {nombre_archivo}")

        # varios archivos
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for nombre, archivo in archivos:
                
                if isinstance(archivo, BytesIO):
                    archivo.seek(0)
                    zipf.writestr(nombre, archivo.read())
                else:
                    zipf.writestr(nombre, archivo)

        zip_buffer.seek(0)

        return zip_buffer
    
    except Exception as e:

        print("=========== ERROR EN GENERAR_EXCEL ===========")
        print(traceback.format_exc())

        raise e

def obtener_datos_dashboard():

    datos = obtener_datos()

    datos_livianos = []

    for item in datos:

        datos_livianos.append({
            "ec5_uuid": item.get("ec5_uuid"),
            "title": item.get("title"),
            "created_at": item.get("created_at"),
            "7_CIUDAD": item.get("7_CIUDAD"),
            "8_DEPARTAMENTO": item.get("8_DEPARTAMENTO"),
            "66_CAPACIDAD_UPS_KVA": item.get("66_CAPACIDAD_UPS_KVA"),
            "69_NUMERO_DE_SERIE_D": item.get("69_NUMERO_DE_SERIE_D"),
            "11_CODIGO_TECNICO": item.get("11_CODIGO_TECNICO"),
            "5_NOMBRE_SEDE": item.get("5_NOMBRE_SEDE"),
            "6_DIRECCION": item.get("6_DIRECCION"),
            "2_ID_SEDE": item.get("2_ID_SEDE")
        })

    return datos_livianos

# ============================
# RUTAS DEL MODULO SENA
# ============================

def registrar_rutas(app):

    @app.route("/sena/datos", methods=["GET"])
    def datos():

        return jsonify(obtener_datos_dashboard())


    @app.route("/sena/generar", methods=["POST"])
    def generar():

        try:

            data = request.json
            seleccionados = data.get("ids", [])

            archivo = generar_excel(seleccionados)

            if not archivo:
                return {"error": "No se generaron archivos"}, 500

            if isinstance(archivo, tuple):

                nombre, buffer = archivo

                return send_file(
                    BytesIO(buffer),
                    as_attachment=True,
                    download_name=nombre,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            return send_file(
                archivo,
                as_attachment=True,
                download_name="informes.zip",
                mimetype="application/zip"
            )

        except Exception as e:

            print(traceback.format_exc())

            return {"error": str(e)}, 500